#!/usr/bin/env python3
"""
Import des clients depuis Excel vers Odoo - AH CHOU
"""

import logging
from openpyxl import load_workbook
from odoo_client import OdooClient
from config import (
    EXCEL_FILE, IMPORT_LIMIT, DRY_RUN,
    GROUPE_CLIENT_MAPPING, ENSEIGNE_MAPPING, 
    TYPOLOGIE_MAPPING, CATEGORIE_CLIENT_MAPPING,
    CONDITION_PAIEMENT_MAPPING
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ClientImporter:
    """Importeur de clients depuis Excel vers Odoo"""
    
    SHEET_NAME = 'trame client'
    
    # Mapping colonnes Excel -> index (0-based)
    COLUMNS = {
        'code_interne': 0,        # Code interne
        'nom_client': 1,          # Nom Client
        'raison_sociale': 2,      # Raison sociale
        'telephone': 3,           # Téléphone
        'telephone2': 4,          # Téléphone 2
        'fax': 5,                 # Fax
        'gsm': 6,                 # GSM
        'civilite_1': 7,          # Civilité 1
        'responsable_1': 8,       # Responsable 1
        'civilite_2': 9,          # Civilité 2
        'responsable_2': 10,      # Responsable 2
        'num_gr_clients': 11,     # N° Gr clients
        'nom_gr_clients': 12,     # Nom Gr. clients
        'num_enseigne': 13,       # N° Enseig.
        'nom_enseigne': 14,       # Nom Enseigne
        'num_categorie': 15,      # N° Catégorie
        'nom_categorie': 16,      # Nom catégorie
        'base_tarif': 17,         # Base tarif
        'plafond_encours': 18,    # Plafond d'encours
        'num_poste': 19,          # N° Poste
        'commentaires': 20,       # Commentaires
        'code_reglement': 21,     # Code règlement
        'code_tarif': 22,         # Code tarif
        'abonne': 23,             # Abonné
        'prix_net': 24,           # Prix Net
        'date_debut_msg': 25,     # Date début message
        'date_fin_msg': 26,       # Date de fin de message
        'message': 27,            # Message
        'code_pays': 28,          # Code pays
        'code_devise': 29,        # Code devise
        'national': 30,           # National
        'email': 31,              # E-Mail
        'site_web': 32,           # Site Web
        'tx_remise': 33,          # Tx remise
        'tx_rfa': 34,             # Tx RFA
        'budget_pptg': 35,        # Budget PPTG
        'taux_pptg': 36,          # Taux PPTG
        'contrat': 37,            # Contrat
        'maj_ref': 38,            # MAJ Réf.
        'num_compta': 39,         # N° Compta
        'num_compte_vente': 40,   # N° compte vente
        'num_compte_achat': 41,   # N° compte achat
        'siret': 42,              # Siret
        'code_ape': 43,           # Code APE
        'risque': 44,             # Risque
        'blocage': 45,            # Blocage
        'groupe_fact': 46,        # Groupe Fact.
        'resp_compta': 47,        # Resp. Cpta.
        'tel_compta': 48,         # Tél. Cpta.
        'gsm_compta': 49,         # GSM Cpta.
        'fax_compta': 50,         # Fax Cpta
        'email_compta': 51,       # Email Cpta.
        'client_ref': 52,         # Client de réf.
        'infos_facture': 53,      # Infos Facture
        'vente_au_pr': 54,        # Vente au PR
        'tx_sur_prht': 55,        # Tx sur PRHT
        'infos_legales': 56,      # Infos légales
    }
    
    def __init__(self):
        self.odoo = OdooClient()
        self.stats = {
            'total': 0,
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': 0,
        }
        self.category_cache = {}
        self.pricelist_cache = {}
        self.payment_term_cache = {}
    
    def get_cell_value(self, row, col_name):
        """Récupère la valeur d'une cellule"""
        idx = self.COLUMNS.get(col_name)
        if idx is None or idx >= len(row):
            return None
        value = row[idx]
        if value is None:
            return None
        if isinstance(value, str):
            return value.strip() if value.strip() else None
        return value
    
    def get_category_id(self, category_name):
        """Récupère ou crée une catégorie client"""
        if not category_name:
            return None
        
        # Normaliser le nom
        normalized = CATEGORIE_CLIENT_MAPPING.get(category_name.upper(), category_name)
        
        if normalized in self.category_cache:
            return self.category_cache[normalized]
        
        cat_id = self.odoo.get_or_create_category(normalized)
        self.category_cache[normalized] = cat_id
        return cat_id
    
    def get_pricelist_id(self, tarif_code):
        """Récupère la liste de prix selon le code tarif"""
        if not tarif_code:
            return None
        
        # Mapping code tarif -> nom liste de prix
        tarif_mapping = {
            '00': 'Tarif Base',
            '01': 'Tarif T1',
            '02': 'Tarif T2',
            '03': 'Tarif T3',
            '04': 'Tarif T4',
            '05': 'Tarif T5',
            '06': 'Tarif T6',
        }
        
        tarif_name = tarif_mapping.get(str(tarif_code).zfill(2), 'Tarif Base')
        
        if tarif_name in self.pricelist_cache:
            return self.pricelist_cache[tarif_name]
        
        pl_id = self.odoo.get_pricelist_id(tarif_name)
        self.pricelist_cache[tarif_name] = pl_id
        return pl_id
    
    def get_payment_term_id(self, code_reglement):
        """Récupère la condition de paiement"""
        if not code_reglement:
            return None
        
        term_name = CONDITION_PAIEMENT_MAPPING.get(str(code_reglement).zfill(2))
        if not term_name:
            return None
        
        if term_name in self.payment_term_cache:
            return self.payment_term_cache[term_name]
        
        pt_id = self.odoo.get_payment_term_id(term_name)
        self.payment_term_cache[term_name] = pt_id
        return pt_id
    
    def normalize_groupe_client(self, value):
        """Normalise le groupe client"""
        if not value:
            return None
        return GROUPE_CLIENT_MAPPING.get(value.upper(), 'autres')
    
    def normalize_enseigne(self, value):
        """Normalise l'enseigne"""
        if not value:
            return None
        return ENSEIGNE_MAPPING.get(value.upper(), 'independant')
    
    def normalize_typologie(self, value):
        """Normalise la typologie"""
        if not value:
            return None
        return TYPOLOGIE_MAPPING.get(value.upper(), 'autres')
    
    def normalize_blocage(self, value):
        """Normalise le type de blocage"""
        if not value:
            return 'non'
        value_upper = str(value).upper()
        if 'OUI' in value_upper and 'DEPASSEMENT' in value_upper:
            return 'oui_depassement'
        elif 'OUI' in value_upper:
            return 'oui_toujours'
        return 'non'
    
    def row_to_partner_values(self, row):
        """Convertit une ligne Excel en valeurs res.partner"""
        
        nom = self.get_cell_value(row, 'nom_client')
        if not nom:
            return None
        
        # Valeurs de base
        values = {
            'name': nom,
            'is_company': True,
            'customer_rank': 1,
            'ref': self.get_cell_value(row, 'code_interne'),
        }
        
        # Contact
        phone = self.get_cell_value(row, 'telephone')
        if phone:
            values['phone'] = str(phone)
        
        mobile = self.get_cell_value(row, 'gsm')
        if mobile:
            values['mobile'] = str(mobile)
        
        email = self.get_cell_value(row, 'email')
        if email:
            values['email'] = email
        
        website = self.get_cell_value(row, 'site_web')
        if website:
            values['website'] = website
        
        # Identification
        siret = self.get_cell_value(row, 'siret')
        if siret:
            values['siret'] = str(siret).replace(' ', '')
        
        # Finances
        plafond = self.get_cell_value(row, 'plafond_encours')
        if plafond:
            try:
                values['credit_limit'] = float(plafond)
            except (ValueError, TypeError):
                pass
        
        # Catégorie
        cat_name = self.get_cell_value(row, 'nom_categorie')
        if cat_name:
            cat_id = self.get_category_id(cat_name)
            if cat_id:
                values['category_id'] = [(6, 0, [cat_id])]
        
        # Liste de prix
        code_tarif = self.get_cell_value(row, 'code_tarif')
        pl_id = self.get_pricelist_id(code_tarif)
        if pl_id:
            values['property_product_pricelist'] = pl_id
        
        # Condition de paiement
        code_reglement = self.get_cell_value(row, 'code_reglement')
        pt_id = self.get_payment_term_id(code_reglement)
        if pt_id:
            values['property_payment_term_id'] = pt_id
        
        # Champs personnalisés (x_)
        # Note: Ces champs doivent être créés dans Odoo avant l'import
        custom_fields = {
            'x_groupe_client': self.normalize_groupe_client(
                self.get_cell_value(row, 'nom_gr_clients')
            ),
            'x_enseigne': self.normalize_enseigne(
                self.get_cell_value(row, 'nom_enseigne')
            ),
            'x_taux_remise': self.get_cell_value(row, 'tx_remise'),
            'x_taux_rfa': self.get_cell_value(row, 'tx_rfa'),
            'x_blocage': self.normalize_blocage(
                self.get_cell_value(row, 'blocage')
            ),
            'x_compte_comptable': self.get_cell_value(row, 'num_compta'),
            'x_responsable_1': self.get_cell_value(row, 'responsable_1'),
            'x_responsable_2': self.get_cell_value(row, 'responsable_2'),
            'x_resp_compta': self.get_cell_value(row, 'resp_compta'),
            'x_tel_compta': self.get_cell_value(row, 'tel_compta'),
            'x_email_compta': self.get_cell_value(row, 'email_compta'),
            'x_groupe_facturation': self.get_cell_value(row, 'groupe_fact'),
        }
        
        # Ajouter seulement les champs non-None
        for key, val in custom_fields.items():
            if val is not None:
                values[key] = val
        
        return values
    
    def import_clients(self):
        """Importe tous les clients depuis Excel"""
        logger.info(f"Chargement du fichier Excel: {EXCEL_FILE}")
        
        wb = load_workbook(EXCEL_FILE, read_only=True)
        
        if self.SHEET_NAME not in wb.sheetnames:
            logger.error(f"Feuille '{self.SHEET_NAME}' non trouvée")
            return
        
        ws = wb[self.SHEET_NAME]
        rows = list(ws.iter_rows(values_only=True))
        
        # Skip header (première ligne)
        data_rows = rows[1:]
        
        if IMPORT_LIMIT:
            data_rows = data_rows[:IMPORT_LIMIT]
        
        self.stats['total'] = len(data_rows)
        logger.info(f"Import de {self.stats['total']} clients...")
        
        for i, row in enumerate(data_rows, 1):
            try:
                values = self.row_to_partner_values(row)
                
                if not values:
                    self.stats['skipped'] += 1
                    continue
                
                # Chercher par référence ou nom
                ref = values.get('ref')
                domain = []
                if ref:
                    domain = [('ref', '=', ref)]
                else:
                    domain = [('name', '=', values['name'])]
                
                existing = self.odoo.search('res.partner', domain, limit=1)
                
                if existing:
                    # Mise à jour
                    self.odoo.write('res.partner', existing, values)
                    self.stats['updated'] += 1
                    logger.debug(f"[{i}] MAJ: {values['name']}")
                else:
                    # Création
                    self.odoo.create('res.partner', values)
                    self.stats['created'] += 1
                    logger.debug(f"[{i}] Créé: {values['name']}")
                
                # Progress
                if i % 50 == 0:
                    logger.info(f"Progression: {i}/{self.stats['total']}")
                    
            except Exception as e:
                self.stats['errors'] += 1
                logger.error(f"Erreur ligne {i}: {e}")
        
        wb.close()
        self.print_stats()
    
    def print_stats(self):
        """Affiche les statistiques d'import"""
        print("\n" + "="*50)
        print("RÉSULTATS IMPORT CLIENTS")
        print("="*50)
        print(f"  Total lignes     : {self.stats['total']}")
        print(f"  Créés            : {self.stats['created']}")
        print(f"  Mis à jour       : {self.stats['updated']}")
        print(f"  Ignorés          : {self.stats['skipped']}")
        print(f"  Erreurs          : {self.stats['errors']}")
        print("="*50)
        if DRY_RUN:
            print("⚠️  MODE DRY-RUN: Aucune modification réelle")


def main():
    """Point d'entrée principal"""
    print("="*50)
    print("IMPORT CLIENTS - AH CHOU vers ODOO")
    print("="*50)
    
    if DRY_RUN:
        print("⚠️  Mode DRY-RUN activé (pas d'écriture)")
    
    importer = ClientImporter()
    importer.import_clients()


if __name__ == '__main__':
    main()
