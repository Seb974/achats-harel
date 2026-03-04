#!/usr/bin/env python3
"""
SCRIPT D'IMPORT DES DONNÉES VERS ODOO
=====================================
GEL OI / AH CHOU - Bloc 6 Import des données

Usage:
    python import_odoo.py --type fournisseurs
    python import_odoo.py --type clients
    python import_odoo.py --type produits
    python import_odoo.py --type all
    
Options:
    --dry-run    Simulation sans écriture
    --limit N    Limiter à N enregistrements
"""

import xmlrpc.client
import argparse
import sys
import os
from openpyxl import load_workbook
from datetime import datetime

# Configuration Odoo
ODOO_URL = "https://ah-chou1.odoo.com"
ODOO_DB = "ah-chou1"
ODOO_USER = "mathieu.loic.hoarau@gmail.com"
ODOO_PASS = "gbtN0WxuCVjg@g*C"

# Chemins des fichiers
BASE_PATH = "/Users/mhoar/Desktop/GEL OI/donnees"
FILES = {
    'fournisseurs': os.path.join(BASE_PATH, "9-Liste des fournisseurs.xlsx"),
    'clients': os.path.join(BASE_PATH, "Feuille de calcul sans titre (3).xlsx"),
    'produits': os.path.join(BASE_PATH, "7-Liste des Produits.xlsx"),
}

# ==============================================================================
# CONNEXION ODOO
# ==============================================================================

def connect_odoo():
    """Établit la connexion à Odoo"""
    common = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/common')
    uid = common.authenticate(ODOO_DB, ODOO_USER, ODOO_PASS, {})
    if not uid:
        raise Exception("Authentification Odoo échouée")
    models = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object')
    return uid, models

def execute(models, uid, model, method, *args, **kwargs):
    """Exécute une méthode sur un modèle Odoo"""
    return models.execute_kw(ODOO_DB, uid, ODOO_PASS, model, method, *args, **kwargs)

# ==============================================================================
# UTILITAIRES
# ==============================================================================

def clean_value(value):
    """Nettoie une valeur Excel"""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if value == '' or value.lower() in ['n/a', 'na', '-', 'none', '0']:
            return None
    return value

def parse_float(value):
    """Parse une valeur flottante"""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        value = str(value).replace(',', '.').replace(' ', '').replace('€', '').replace('%', '')
        return float(value)
    except:
        return 0.0

def parse_int(value):
    """Parse une valeur entière"""
    if value is None:
        return 0
    try:
        return int(float(value))
    except:
        return 0

def parse_boolean(value):
    """Parse une valeur booléenne"""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    str_val = str(value).lower().strip()
    return str_val in ['oui', 'yes', 'true', '1', 'x', 'o']

def read_excel(filepath, sheet_name=None):
    """Lit un fichier Excel et retourne les données"""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h else f'col_{i}' for i, h in enumerate(rows[0])]
    
    data = []
    for row in rows[1:]:
        if any(cell is not None for cell in row):
            record = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
            data.append(record)
    
    wb.close()
    return data

# ==============================================================================
# IMPORT FOURNISSEURS
# ==============================================================================

def import_fournisseurs(uid, models, dry_run=False, limit=None):
    """Importe les fournisseurs dans Odoo"""
    print("\n" + "=" * 70)
    print("IMPORT FOURNISSEURS")
    print("=" * 70)
    
    data = read_excel(FILES['fournisseurs'], 'A')
    if limit:
        data = data[:limit]
    
    print(f"📄 {len(data)} fournisseurs à importer")
    
    created = 0
    updated = 0
    errors = []
    
    for i, row in enumerate(data):
        try:
            code = clean_value(row.get('Code fournisseur'))
            name = clean_value(row.get('Nom fournisseur'))
            
            if not name:
                continue
            
            # Préparer les données
            partner_data = {
                'name': name,
                'supplier_rank': 1,
                'is_company': True,
                'country_id': 75,  # France
                'lang': 'fr_FR',
            }
            
            if code:
                partner_data['ref'] = code
            
            # Champs optionnels
            if clean_value(row.get('Adresse')):
                partner_data['street'] = row['Adresse']
            if clean_value(row.get('Code postal')):
                partner_data['zip'] = str(row['Code postal'])
            if clean_value(row.get('Ville')):
                partner_data['city'] = row['Ville']
            if clean_value(row.get('Téléphone')):
                partner_data['phone'] = str(row['Téléphone'])
            if clean_value(row.get('E-Mail')):
                partner_data['email'] = row['E-Mail']
            if clean_value(row.get('Site Web')):
                partner_data['website'] = row['Site Web']
            if clean_value(row.get('Commentaires')):
                partner_data['comment'] = row['Commentaires']
            
            # Champs personnalisés
            if clean_value(row.get('Responsable 1')):
                partner_data['x_responsable_1'] = row['Responsable 1']
            if clean_value(row.get('Responsable 2')):
                partner_data['x_responsable_2'] = row['Responsable 2']
            if clean_value(row.get('N° Client')):
                partner_data['x_numero_client_chez_frs'] = str(row['N° Client'])
            if clean_value(row.get('N° Cpte. Cpta.')):
                partner_data['x_compte_comptable'] = str(row['N° Cpte. Cpta.'])
            if clean_value(row.get('C. Cpta. Aux.')):
                partner_data['x_compte_auxiliaire'] = str(row['C. Cpta. Aux.'])
            if clean_value(row.get('Condi. livraison')):
                partner_data['x_incoterm'] = str(row['Condi. livraison'])
            if clean_value(row.get('Condi. achats')):
                partner_data['x_conditions_achat'] = str(row['Condi. achats'])
            if row.get('Nb Jrs Dispo'):
                partner_data['x_delai_livraison'] = parse_int(row['Nb Jrs Dispo'])
            
            # Transitaire
            if parse_boolean(row.get('Transitaire')):
                partner_data['x_is_transitaire'] = True
            
            if dry_run:
                print(f"   [DRY-RUN] {name}")
                created += 1
                continue
            
            # Vérifier si existe déjà
            existing = execute(models, uid, 'res.partner', 'search', 
                [[('ref', '=', code), ('supplier_rank', '>', 0)]]) if code else []
            
            if existing:
                execute(models, uid, 'res.partner', 'write', [existing, partner_data])
                updated += 1
                print(f"   ✏️  [{code}] {name} (mis à jour)")
            else:
                partner_id = execute(models, uid, 'res.partner', 'create', [partner_data])
                created += 1
                print(f"   ✅ [{code}] {name} (ID: {partner_id})")
                
        except Exception as e:
            errors.append(f"Ligne {i+2}: {e}")
            print(f"   ❌ Erreur ligne {i+2}: {e}")
    
    print(f"\n📊 Résumé: {created} créés, {updated} mis à jour, {len(errors)} erreurs")
    return created, updated, errors

# ==============================================================================
# IMPORT CLIENTS
# ==============================================================================

def import_clients(uid, models, dry_run=False, limit=None):
    """Importe les clients dans Odoo"""
    print("\n" + "=" * 70)
    print("IMPORT CLIENTS")
    print("=" * 70)
    
    data = read_excel(FILES['clients'], 'CLIENT')
    if limit:
        data = data[:limit]
    
    print(f"📄 {len(data)} clients à importer")
    
    # Mapping tarifs
    TARIFS = {'00': 2, '90': 3, '91': 4, '92': 5, '93': 6}
    
    created = 0
    updated = 0
    errors = []
    
    for i, row in enumerate(data):
        try:
            code = clean_value(row.get('Code interne'))
            name = clean_value(row.get('Nom Client'))
            
            if not name:
                continue
            
            # Préparer les données
            partner_data = {
                'name': name,
                'customer_rank': 1,
                'is_company': True,
                'country_id': 75,  # France
                'lang': 'fr_FR',
            }
            
            if code:
                partner_data['ref'] = code
            
            # Champs optionnels
            # Note: company_name n'existe pas dans Odoo 19, on ignore la raison sociale
            if clean_value(row.get('Téléphone')):
                partner_data['phone'] = str(row['Téléphone'])
            # Note: mobile n'existe pas dans Odoo 19
            if clean_value(row.get('E-Mail')):
                partner_data['email'] = row['E-Mail']
            
            # Liste de prix
            code_tarif = clean_value(row.get('Code tarif'))
            if code_tarif and str(code_tarif) in TARIFS:
                partner_data['property_product_pricelist'] = TARIFS[str(code_tarif)]
            
            # Champs personnalisés
            if clean_value(row.get('Responsable 1')):
                partner_data['x_responsable_1'] = row['Responsable 1']
            if clean_value(row.get('Responsable 2')):
                partner_data['x_responsable_2'] = row['Responsable 2']
            if row.get('Tx remise'):
                partner_data['x_taux_remise'] = parse_float(row['Tx remise'])
            if row.get('Tx RFA'):
                partner_data['x_taux_rfa'] = parse_float(row['Tx RFA'])
            if clean_value(row.get('N° Compta')):
                partner_data['x_compte_comptable'] = str(row['N° Compta'])
            if clean_value(row.get('Groupe Fact.')):
                partner_data['x_groupe_facturation'] = str(row['Groupe Fact.'])
            if clean_value(row.get('Resp. Cpta.')):
                partner_data['x_resp_compta'] = str(row['Resp. Cpta.'])
            if clean_value(row.get('Tél. Cpta.')):
                partner_data['x_tel_compta'] = str(row['Tél. Cpta.'])
            if clean_value(row.get('Email Cpta.')):
                partner_data['x_email_compta'] = str(row['Email Cpta.'])
            
            if dry_run:
                print(f"   [DRY-RUN] {name}")
                created += 1
                continue
            
            # Vérifier si existe déjà
            existing = execute(models, uid, 'res.partner', 'search', 
                [[('ref', '=', code), ('customer_rank', '>', 0)]]) if code else []
            
            if existing:
                execute(models, uid, 'res.partner', 'write', [existing, partner_data])
                updated += 1
                print(f"   ✏️  [{code}] {name} (mis à jour)")
            else:
                partner_id = execute(models, uid, 'res.partner', 'create', [partner_data])
                created += 1
                print(f"   ✅ [{code}] {name} (ID: {partner_id})")
                
        except Exception as e:
            errors.append(f"Ligne {i+2}: {e}")
            print(f"   ❌ Erreur ligne {i+2}: {e}")
    
    print(f"\n📊 Résumé: {created} créés, {updated} mis à jour, {len(errors)} erreurs")
    return created, updated, errors

# ==============================================================================
# IMPORT PRODUITS
# ==============================================================================

def import_produits(uid, models, dry_run=False, limit=None):
    """Importe les produits dans Odoo"""
    print("\n" + "=" * 70)
    print("IMPORT PRODUITS")
    print("=" * 70)
    
    data = read_excel(FILES['produits'], 'A')
    if limit:
        data = data[:limit]
    
    print(f"📄 {len(data)} produits à importer")
    
    created = 0
    updated = 0
    errors = []
    
    for i, row in enumerate(data):
        try:
            code = clean_value(row.get('N° Interne'))
            name = clean_value(row.get('Désignation'))
            
            if not name:
                continue
            
            # Préparer les données
            product_data = {
                'name': name,
                'type': 'consu',             # Goods (Odoo 19 = 'consu' pour stockables)
                'sale_ok': True,
                'purchase_ok': True,
                'categ_id': 4,               # GEL OI par défaut
            }
            
            if code:
                product_data['default_code'] = str(code)
            
            # Code-barres
            ean = clean_value(row.get('Ean 13'))
            if ean and len(str(ean)) >= 8:
                product_data['barcode'] = str(ean)
            
            # Prix
            if row.get('PV HT'):
                product_data['list_price'] = parse_float(row['PV HT'])
            if row.get('PR HT'):
                product_data['standard_price'] = parse_float(row['PR HT'])
            
            # Poids
            if row.get('Poids net'):
                product_data['weight'] = parse_float(row['Poids net'])
            
            # Champs personnalisés
            if clean_value(row.get('Nom long')):
                product_data['x_nom_long'] = row['Nom long']
            if row.get('Contenu'):
                product_data['x_contenu'] = parse_float(row['Contenu'])
            if row.get('Condt.'):
                product_data['x_conditionnement'] = parse_float(row['Condt.'])
            if row.get('PCB'):
                product_data['x_pcb'] = parse_float(row['PCB'])
            if row.get('Poids brut'):
                product_data['x_poids_brut'] = parse_float(row['Poids brut'])
            if row.get('Coef. App.'):
                product_data['x_coef_approche'] = parse_float(row['Coef. App.'])
            if row.get('PRMUP'):
                product_data['x_prmup'] = parse_float(row['PRMUP'])
            
            # Tarifs T1-T6
            for t in range(1, 7):
                key = f'T{t} HT'
                if row.get(key):
                    product_data[f'x_tarif_t{t}_ht'] = parse_float(row[key])
            
            # Marque et Origine
            if clean_value(row.get('Marque')):
                product_data['x_marque'] = row['Marque']
            if clean_value(row.get('Origine')):
                product_data['x_origine'] = row['Origine']
            
            # EAN dans champ personnalisé aussi
            if ean:
                product_data['x_gencod'] = str(ean)
            
            if dry_run:
                print(f"   [DRY-RUN] [{code}] {name}")
                created += 1
                continue
            
            # Vérifier si existe déjà
            existing = execute(models, uid, 'product.template', 'search', 
                [[('default_code', '=', code)]]) if code else []
            
            if existing:
                execute(models, uid, 'product.template', 'write', [existing, product_data])
                updated += 1
                print(f"   ✏️  [{code}] {name[:40]} (mis à jour)")
            else:
                product_id = execute(models, uid, 'product.template', 'create', [product_data])
                created += 1
                print(f"   ✅ [{code}] {name[:40]} (ID: {product_id})")
                
        except Exception as e:
            errors.append(f"Ligne {i+2}: {e}")
            print(f"   ❌ Erreur ligne {i+2}: {e}")
    
    print(f"\n📊 Résumé: {created} créés, {updated} mis à jour, {len(errors)} erreurs")
    return created, updated, errors

# ==============================================================================
# MAIN
# ==============================================================================

def main():
    parser = argparse.ArgumentParser(description='Import données vers Odoo')
    parser.add_argument('--type', choices=['fournisseurs', 'clients', 'produits', 'all'],
                        required=True, help='Type de données à importer')
    parser.add_argument('--dry-run', action='store_true', 
                        help='Simulation sans écriture')
    parser.add_argument('--limit', type=int, default=None,
                        help='Limiter à N enregistrements')
    
    args = parser.parse_args()
    
    print("=" * 70)
    print(f"IMPORT ODOO - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)
    
    if args.dry_run:
        print("🔶 MODE SIMULATION (dry-run)")
    
    # Connexion
    print("\n🔌 Connexion à Odoo...")
    uid, models = connect_odoo()
    print(f"   ✅ Connecté (UID: {uid})")
    
    # Import
    results = {}
    
    if args.type in ['fournisseurs', 'all']:
        results['fournisseurs'] = import_fournisseurs(uid, models, args.dry_run, args.limit)
    
    if args.type in ['clients', 'all']:
        results['clients'] = import_clients(uid, models, args.dry_run, args.limit)
    
    if args.type in ['produits', 'all']:
        results['produits'] = import_produits(uid, models, args.dry_run, args.limit)
    
    # Résumé final
    print("\n" + "=" * 70)
    print("RÉSUMÉ FINAL")
    print("=" * 70)
    
    total_created = 0
    total_updated = 0
    total_errors = 0
    
    for data_type, (created, updated, errors) in results.items():
        print(f"\n{data_type.upper()}:")
        print(f"   ✅ Créés: {created}")
        print(f"   ✏️  Mis à jour: {updated}")
        print(f"   ❌ Erreurs: {len(errors)}")
        total_created += created
        total_updated += updated
        total_errors += len(errors)
    
    print(f"\n{'='*70}")
    print(f"TOTAL: {total_created} créés, {total_updated} mis à jour, {total_errors} erreurs")
    print("=" * 70)

if __name__ == '__main__':
    main()
