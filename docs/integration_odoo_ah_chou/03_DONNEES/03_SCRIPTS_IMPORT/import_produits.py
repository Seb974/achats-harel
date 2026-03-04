#!/usr/bin/env python3
"""
Import des produits depuis Excel vers Odoo - AH CHOU
"""

import logging
from openpyxl import load_workbook
from odoo_client import OdooClient
from config import EXCEL_FILE, IMPORT_LIMIT, DRY_RUN, TVA_MAPPING

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ProductImporter:
    """Importeur de produits depuis Excel vers Odoo"""
    
    SHEET_NAME = 'Trame articles'
    
    # Mapping colonnes principales (index 0-based)
    COLUMNS = {
        'code_puit': 0,           # C. Puit
        'num_interne': 1,         # N° Interne
        'autre_code': 2,          # Autre C. Interne
        'ean13': 3,               # Ean 13
        'designation': 4,         # Désignation
        'nom_long': 5,            # Nom long
        'argu_recherche': 6,      # Argu. Recherche
        'contenu': 7,             # Contenu
        'condt': 8,               # Condt.
        'pcb': 9,                 # PCB
        'detail_cdt': 10,         # Détail Cdt.
        'poids_brut': 11,         # Poids brut
        'poids_net': 12,          # Poids net
        'unite_poids': 13,        # Unité poids
        'volume': 14,             # Volume
        'unite_volume': 15,       # Unité volume
        'pds_vol': 16,            # Pds-Vol
        'ct': 17,                 # CT
        'tx_tva': 18,             # Tx. TVA
        'pv_ht': 19,              # PV HT
        'pv_ttc': 20,             # PV TTC
        'pr_ht': 21,              # PR HT
        'pr_ttc': 22,             # PR TTC
        'prmup': 23,              # PRMUP
        'tx_com': 24,             # Tx Com.
        # ... colonnes intermédiaires
        'coef_app': 34,           # Coef. App.
        't1_ht': 35,              # T1 HT
        't2_ht': 36,              # T2 HT
        't3_ht': 37,              # T3 HT
        't4_ht': 38,              # T4 HT
        't5_ht': 39,              # T5 HT
        't6_ht': 40,              # T6 HT
        # ... colonnes intermédiaires
        'code_onu': 47,           # Code ONU
        'nom_onu': 48,            # Nom ONU
        'details_nsa': 49,        # Détails NSA (nom scientifique)
        # ... colonnes intermédiaires
        'tx_om': 56,              # Tx OM
        # ... colonnes intermédiaires
        'marque': 61,             # Marque
        'origine': 62,            # Origine
        'c_rayon': 63,            # C. Rayon
        'rayon': 64,              # Rayon
        'gr_art': 65,             # Gr. Art
        'nom_gr': 66,             # Nom GR
        'sgr_art': 67,            # SGr Art
        'nom_sgr': 68,            # Nom SGR
        'segments': 69,           # Segments
        'nom_seg': 70,            # Nom Seg.
    }
    
    def __init__(self):
        self.odoo = OdooClient()
        self.stats = {
            'total': 0,
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': 0,
        }
        self.category_cache = {}
        self.tax_cache = {}
        self.uom_cache = {}
    
    def get_cell_value(self, row, col_name):
        """Récupère la valeur d'une cellule"""
        idx = self.COLUMNS.get(col_name)
        if idx is None or idx >= len(row):
            return None
        value = row[idx]
        if value is None:
            return None
        if isinstance(value, str):
            return value.strip() if value.strip() else None
        return value
    
    def get_float(self, row, col_name, default=0.0):
        """Récupère une valeur float"""
        val = self.get_cell_value(row, col_name)
        if val is None:
            return default
        try:
            return float(val)
        except (ValueError, TypeError):
            return default
    
    def get_category_id(self, rayon, groupe, sous_groupe):
        """Récupère ou crée la hiérarchie de catégories produit"""
        
        # Niveau 1: Rayon
        if not rayon:
            rayon = 'Divers'
        
        rayon_key = f"rayon:{rayon}"
        if rayon_key not in self.category_cache:
            rayon_id = self.odoo.get_or_create_product_category(rayon)
            self.category_cache[rayon_key] = rayon_id
        rayon_id = self.category_cache[rayon_key]
        
        # Niveau 2: Groupe
        if groupe:
            groupe_key = f"groupe:{rayon}:{groupe}"
            if groupe_key not in self.category_cache:
                groupe_id = self.odoo.get_or_create_product_category(groupe, rayon_id)
                self.category_cache[groupe_key] = groupe_id
            groupe_id = self.category_cache[groupe_key]
            
            # Niveau 3: Sous-groupe
            if sous_groupe:
                sg_key = f"sgroupe:{rayon}:{groupe}:{sous_groupe}"
                if sg_key not in self.category_cache:
                    sg_id = self.odoo.get_or_create_product_category(sous_groupe, groupe_id)
                    self.category_cache[sg_key] = sg_id
                return self.category_cache[sg_key]
            
            return groupe_id
        
        return rayon_id
    
    def get_tax_ids(self, taux_tva):
        """Récupère les IDs des taxes selon le taux"""
        if taux_tva is None:
            return []
        
        try:
            taux = float(taux_tva)
        except (ValueError, TypeError):
            return []
        
        tax_name = TVA_MAPPING.get(taux)
        if not tax_name:
            return []
        
        if tax_name in self.tax_cache:
            tax_id = self.tax_cache[tax_name]
        else:
            tax_id = self.odoo.get_tax_id(tax_name)
            self.tax_cache[tax_name] = tax_id
        
        return [tax_id] if tax_id else []
    
    def get_uom_id(self, uom_name='Unité(s)'):
        """Récupère l'ID de l'unité de mesure"""
        if uom_name in self.uom_cache:
            return self.uom_cache[uom_name]
        
        uom = self.odoo.search_read(
            'uom.uom',
            [('name', 'ilike', uom_name)],
            ['id'],
            limit=1
        )
        
        uom_id = uom[0]['id'] if uom else None
        self.uom_cache[uom_name] = uom_id
        return uom_id
    
    def row_to_product_values(self, row):
        """Convertit une ligne Excel en valeurs product.template"""
        
        designation = self.get_cell_value(row, 'designation')
        if not designation:
            return None
        
        code_puit = self.get_cell_value(row, 'code_puit')
        
        # Valeurs de base
        values = {
            'name': designation,
            'default_code': code_puit,
            'type': 'product',  # Produit stockable
            'sale_ok': True,
            'purchase_ok': True,
        }
        
        # Code-barres
        ean13 = self.get_cell_value(row, 'ean13')
        if ean13:
            values['barcode'] = str(ean13)
        
        # Prix
        pv_ht = self.get_float(row, 'pv_ht')
        if pv_ht > 0:
            values['list_price'] = pv_ht
        
        pr_ht = self.get_float(row, 'pr_ht')
        if pr_ht > 0:
            values['standard_price'] = pr_ht
        
        # Poids
        poids_net = self.get_float(row, 'poids_net')
        if poids_net > 0:
            values['weight'] = poids_net
        
        # Volume
        volume = self.get_float(row, 'volume')
        if volume > 0:
            values['volume'] = volume
        
        # Catégorie
        rayon = self.get_cell_value(row, 'rayon')
        groupe = self.get_cell_value(row, 'nom_gr')
        sous_groupe = self.get_cell_value(row, 'nom_sgr')
        categ_id = self.get_category_id(rayon, groupe, sous_groupe)
        if categ_id:
            values['categ_id'] = categ_id
        
        # Taxes
        taux_tva = self.get_cell_value(row, 'tx_tva')
        tax_ids = self.get_tax_ids(taux_tva)
        if tax_ids:
            values['taxes_id'] = [(6, 0, tax_ids)]
            values['supplier_taxes_id'] = [(6, 0, tax_ids)]
        
        # Traçabilité
        values['tracking'] = 'lot'  # Suivi par lot
        
        # Champs personnalisés (x_)
        custom_fields = {
            'x_code_interne': self.get_cell_value(row, 'num_interne'),
            'x_nom_long': self.get_cell_value(row, 'nom_long'),
            'x_contenu': self.get_float(row, 'contenu'),
            'x_conditionnement': self.get_float(row, 'condt'),
            'x_pcb': self.get_float(row, 'pcb'),
            'x_poids_brut': self.get_float(row, 'poids_brut'),
            'x_coef_approche': self.get_float(row, 'coef_app', 1.0),
            'x_zone_peche': self.get_cell_value(row, 'nom_onu'),
            'x_nom_scientifique': self.get_cell_value(row, 'details_nsa'),
            'x_origine': self.get_cell_value(row, 'origine'),
            'x_marque': self.get_cell_value(row, 'marque'),
            'x_segment': self.get_cell_value(row, 'nom_seg'),
            'x_taux_om': self.get_float(row, 'tx_om'),
            'x_pv_ttc': self.get_float(row, 'pv_ttc'),
            'x_pr_ttc': self.get_float(row, 'pr_ttc'),
            'x_prmup': self.get_float(row, 'prmup'),
            # Tarifs échelonnés
            'x_tarif_t1_ht': self.get_float(row, 't1_ht'),
            'x_tarif_t2_ht': self.get_float(row, 't2_ht'),
            'x_tarif_t3_ht': self.get_float(row, 't3_ht'),
            'x_tarif_t4_ht': self.get_float(row, 't4_ht'),
            'x_tarif_t5_ht': self.get_float(row, 't5_ht'),
            'x_tarif_t6_ht': self.get_float(row, 't6_ht'),
        }
        
        # Ajouter seulement les champs non-None/non-zero pertinents
        for key, val in custom_fields.items():
            if val is not None and val != '' and val != 0:
                values[key] = val
        
        return values
    
    def import_products(self):
        """Importe tous les produits depuis Excel"""
        logger.info(f"Chargement du fichier Excel: {EXCEL_FILE}")
        
        wb = load_workbook(EXCEL_FILE, read_only=True)
        
        if self.SHEET_NAME not in wb.sheetnames:
            logger.error(f"Feuille '{self.SHEET_NAME}' non trouvée")
            return
        
        ws = wb[self.SHEET_NAME]
        rows = list(ws.iter_rows(values_only=True))
        
        # Skip header (première ligne)
        data_rows = rows[1:]
        
        if IMPORT_LIMIT:
            data_rows = data_rows[:IMPORT_LIMIT]
        
        self.stats['total'] = len(data_rows)
        logger.info(f"Import de {self.stats['total']} produits...")
        
        for i, row in enumerate(data_rows, 1):
            try:
                values = self.row_to_product_values(row)
                
                if not values:
                    self.stats['skipped'] += 1
                    continue
                
                # Chercher par référence (default_code)
                ref = values.get('default_code')
                if ref:
                    domain = [('default_code', '=', ref)]
                else:
                    domain = [('name', '=', values['name'])]
                
                existing = self.odoo.search('product.template', domain, limit=1)
                
                if existing:
                    # Mise à jour
                    self.odoo.write('product.template', existing, values)
                    self.stats['updated'] += 1
                    logger.debug(f"[{i}] MAJ: {values['name']}")
                else:
                    # Création
                    self.odoo.create('product.template', values)
                    self.stats['created'] += 1
                    logger.debug(f"[{i}] Créé: {values['name']}")
                
                # Progress
                if i % 50 == 0:
                    logger.info(f"Progression: {i}/{self.stats['total']}")
                    
            except Exception as e:
                self.stats['errors'] += 1
                logger.error(f"Erreur ligne {i}: {e}")
        
        wb.close()
        self.print_stats()
    
    def print_stats(self):
        """Affiche les statistiques d'import"""
        print("\n" + "="*50)
        print("RÉSULTATS IMPORT PRODUITS")
        print("="*50)
        print(f"  Total lignes     : {self.stats['total']}")
        print(f"  Créés            : {self.stats['created']}")
        print(f"  Mis à jour       : {self.stats['updated']}")
        print(f"  Ignorés          : {self.stats['skipped']}")
        print(f"  Erreurs          : {self.stats['errors']}")
        print("="*50)
        if DRY_RUN:
            print("⚠️  MODE DRY-RUN: Aucune modification réelle")


def main():
    """Point d'entrée principal"""
    print("="*50)
    print("IMPORT PRODUITS - AH CHOU vers ODOO")
    print("="*50)
    
    if DRY_RUN:
        print("⚠️  Mode DRY-RUN activé (pas d'écriture)")
    
    importer = ProductImporter()
    importer.import_products()


if __name__ == '__main__':
    main()
