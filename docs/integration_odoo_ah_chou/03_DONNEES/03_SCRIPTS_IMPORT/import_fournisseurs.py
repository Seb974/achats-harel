#!/usr/bin/env python3
"""
Import des fournisseurs depuis Excel vers Odoo - AH CHOU
"""

import logging
from openpyxl import load_workbook
from odoo_client import OdooClient
from config import EXCEL_FILE, IMPORT_LIMIT, DRY_RUN, GROUPE_FOURNISSEUR_MAPPING

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class SupplierImporter:
    """Importeur de fournisseurs depuis Excel vers Odoo"""
    
    SHEET_NAME = 'TRAME FRS'
    
    # Mapping colonnes Excel -> index (0-based)
    COLUMNS = {
        'code_fournisseur': 0,    # Code fournisseur
        'c_interne': 1,           # C. Interne
        'nom_fournisseur': 2,     # Nom fournisseur
        'raison_sociale': 3,      # Raison sociale
        'adresse': 4,             # Adresse
        'code_postal': 5,         # Code postal
        'ville': 6,               # Ville
        'responsable_1': 7,       # Responsable 1
        'responsable_2': 8,       # Responsable 2
        'telephone': 9,           # Téléphone
        'telephone2': 10,         # Téléphone 2
        'fax': 11,                # N° Fax
        'france': 12,             # France (0 ou 1)
        'email': 13,              # E-Mail
        'site_web': 14,           # Site Web
        'grp_frs': 15,            # Grp. Frs
        'num_client': 16,         # N° Client
        'commentaires': 17,       # Commentaires
        'num_cpte_cpta': 18,      # N° Cpte. Cpta.
        'c_cpta_aux': 19,         # C. Cpta. Aux.
        'num_cpte_achat': 20,     # N° Cpte comptable achat
        'transitaire': 21,        # Transitaire
        'nb_jrs_dispo': 22,       # Nb Jrs Dispo
        'condi_paiement': 23,     # Condi. paiement
        'condi_livraison': 24,    # Condi. livraison
        'condi_achats': 25,       # Condi. achats
    }
    
    def __init__(self):
        self.odoo = OdooClient()
        self.stats = {
            'total': 0,
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': 0,
        }
        self.country_cache = {}
    
    def get_cell_value(self, row, col_name):
        """Récupère la valeur d'une cellule"""
        idx = self.COLUMNS.get(col_name)
        if idx is None or idx >= len(row):
            return None
        value = row[idx]
        if value is None:
            return None
        if isinstance(value, str):
            return value.strip() if value.strip() else None
        return value
    
    def get_float(self, row, col_name, default=0.0):
        """Récupère une valeur float"""
        val = self.get_cell_value(row, col_name)
        if val is None:
            return default
        try:
            return float(val)
        except (ValueError, TypeError):
            return default
    
    def get_int(self, row, col_name, default=0):
        """Récupère une valeur int"""
        val = self.get_cell_value(row, col_name)
        if val is None:
            return default
        try:
            return int(val)
        except (ValueError, TypeError):
            return default
    
    def normalize_groupe_fournisseur(self, value):
        """Normalise le groupe fournisseur"""
        if not value:
            return 'autres'
        return GROUPE_FOURNISSEUR_MAPPING.get(value.upper().strip(), 'autres')
    
    def get_country_id(self, is_france, groupe_frs=None, adresse=None):
        """Détermine le pays du fournisseur"""
        
        # Si France = 1, c'est la France
        if is_france == 1 or is_france == 1.0:
            country_name = 'France'
        else:
            # Déduire du groupe fournisseur
            if groupe_frs:
                groupe_upper = groupe_frs.upper()
                if 'INDE' in groupe_upper:
                    country_name = 'India'
                elif 'EUROPE' in groupe_upper:
                    country_name = 'France'  # Par défaut Europe
                elif 'LOCAL' in groupe_upper:
                    country_name = 'France'
                elif 'AFRIQUE' in groupe_upper:
                    country_name = 'South Africa'  # Par défaut
                elif 'AMERIQUE' in groupe_upper:
                    country_name = 'United States'  # Par défaut
                else:
                    country_name = None
            else:
                country_name = None
        
        if not country_name:
            return None
        
        if country_name in self.country_cache:
            return self.country_cache[country_name]
        
        # Rechercher le pays dans Odoo
        country = self.odoo.search_read(
            'res.country',
            [('name', 'ilike', country_name)],
            ['id'],
            limit=1
        )
        
        country_id = country[0]['id'] if country else None
        self.country_cache[country_name] = country_id
        return country_id
    
    def parse_address(self, adresse):
        """Parse une adresse multiligne"""
        if not adresse:
            return None, None, None
        
        lines = str(adresse).strip().split('\n')
        
        if len(lines) == 1:
            return lines[0], None, None
        
        street = lines[0]
        street2 = '\n'.join(lines[1:-1]) if len(lines) > 2 else None
        
        # Dernière ligne peut contenir code postal et ville
        last_line = lines[-1] if len(lines) > 1 else None
        
        return street, street2, None
    
    def row_to_partner_values(self, row):
        """Convertit une ligne Excel en valeurs res.partner (fournisseur)"""
        
        nom = self.get_cell_value(row, 'nom_fournisseur')
        if not nom:
            return None
        
        # Valeurs de base
        values = {
            'name': nom,
            'is_company': True,
            'supplier_rank': 1,  # Marque comme fournisseur
            'customer_rank': 0,
            'ref': self.get_cell_value(row, 'code_fournisseur'),
        }
        
        # Raison sociale
        raison_sociale = self.get_cell_value(row, 'raison_sociale')
        if raison_sociale and raison_sociale != nom:
            values['name'] = raison_sociale
            values['x_nom_commercial'] = nom
        
        # Adresse
        adresse = self.get_cell_value(row, 'adresse')
        if adresse:
            street, street2, _ = self.parse_address(adresse)
            if street:
                values['street'] = street
            if street2:
                values['street2'] = street2
        
        code_postal = self.get_cell_value(row, 'code_postal')
        if code_postal:
            values['zip'] = str(code_postal)
        
        ville = self.get_cell_value(row, 'ville')
        if ville:
            values['city'] = ville
        
        # Pays
        is_france = self.get_float(row, 'france')
        groupe_frs = self.get_cell_value(row, 'grp_frs')
        country_id = self.get_country_id(is_france, groupe_frs, adresse)
        if country_id:
            values['country_id'] = country_id
        
        # Contact
        phone = self.get_cell_value(row, 'telephone')
        if phone:
            values['phone'] = str(phone)
        
        email = self.get_cell_value(row, 'email')
        if email:
            values['email'] = email
        
        website = self.get_cell_value(row, 'site_web')
        if website:
            values['website'] = website
        
        # Champs personnalisés (x_)
        custom_fields = {
            'x_code_interne': self.get_cell_value(row, 'c_interne'),
            'x_groupe_fournisseur': self.normalize_groupe_fournisseur(groupe_frs),
            'x_compte_auxiliaire': self.get_cell_value(row, 'c_cpta_aux'),
            'x_is_transitaire': self.get_float(row, 'transitaire') == 1,
            'x_delai_livraison': self.get_int(row, 'nb_jrs_dispo'),
            'x_incoterm': self.get_cell_value(row, 'condi_livraison'),
            'x_conditions_achat': self.get_cell_value(row, 'condi_achats'),
            'x_responsable_1': self.get_cell_value(row, 'responsable_1'),
            'x_responsable_2': self.get_cell_value(row, 'responsable_2'),
            'x_numero_client_chez_frs': self.get_cell_value(row, 'num_client'),
        }
        
        # Ajouter seulement les champs non-None/non-false pertinents
        for key, val in custom_fields.items():
            if val is not None and val != '' and val != 0 and val is not False:
                values[key] = val
            elif key == 'x_is_transitaire' and val is True:
                values[key] = val
        
        # Commentaires
        commentaires = self.get_cell_value(row, 'commentaires')
        if commentaires:
            values['comment'] = commentaires
        
        return values
    
    def import_suppliers(self):
        """Importe tous les fournisseurs depuis Excel"""
        logger.info(f"Chargement du fichier Excel: {EXCEL_FILE}")
        
        wb = load_workbook(EXCEL_FILE, read_only=True)
        
        if self.SHEET_NAME not in wb.sheetnames:
            logger.error(f"Feuille '{self.SHEET_NAME}' non trouvée")
            return
        
        ws = wb[self.SHEET_NAME]
        rows = list(ws.iter_rows(values_only=True))
        
        # Skip header (première ligne)
        data_rows = rows[1:]
        
        if IMPORT_LIMIT:
            data_rows = data_rows[:IMPORT_LIMIT]
        
        self.stats['total'] = len(data_rows)
        logger.info(f"Import de {self.stats['total']} fournisseurs...")
        
        for i, row in enumerate(data_rows, 1):
            try:
                values = self.row_to_partner_values(row)
                
                if not values:
                    self.stats['skipped'] += 1
                    continue
                
                # Chercher par référence (code fournisseur)
                ref = values.get('ref')
                if ref:
                    domain = [('ref', '=', ref), ('supplier_rank', '>', 0)]
                else:
                    domain = [('name', '=', values['name']), ('supplier_rank', '>', 0)]
                
                existing = self.odoo.search('res.partner', domain, limit=1)
                
                if existing:
                    # Mise à jour
                    self.odoo.write('res.partner', existing, values)
                    self.stats['updated'] += 1
                    logger.debug(f"[{i}] MAJ: {values['name']}")
                else:
                    # Création
                    self.odoo.create('res.partner', values)
                    self.stats['created'] += 1
                    logger.debug(f"[{i}] Créé: {values['name']}")
                
                # Progress
                if i % 100 == 0:
                    logger.info(f"Progression: {i}/{self.stats['total']}")
                    
            except Exception as e:
                self.stats['errors'] += 1
                logger.error(f"Erreur ligne {i}: {e}")
        
        wb.close()
        self.print_stats()
    
    def print_stats(self):
        """Affiche les statistiques d'import"""
        print("\n" + "="*50)
        print("RÉSULTATS IMPORT FOURNISSEURS")
        print("="*50)
        print(f"  Total lignes     : {self.stats['total']}")
        print(f"  Créés            : {self.stats['created']}")
        print(f"  Mis à jour       : {self.stats['updated']}")
        print(f"  Ignorés          : {self.stats['skipped']}")
        print(f"  Erreurs          : {self.stats['errors']}")
        print("="*50)
        if DRY_RUN:
            print("⚠️  MODE DRY-RUN: Aucune modification réelle")


def main():
    """Point d'entrée principal"""
    print("="*50)
    print("IMPORT FOURNISSEURS - AH CHOU vers ODOO")
    print("="*50)
    
    if DRY_RUN:
        print("⚠️  Mode DRY-RUN activé (pas d'écriture)")
    
    importer = SupplierImporter()
    importer.import_suppliers()


if __name__ == '__main__':
    main()
