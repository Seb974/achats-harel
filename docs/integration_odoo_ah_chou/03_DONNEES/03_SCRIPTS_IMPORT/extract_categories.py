#!/usr/bin/env python3
"""
Script d'extraction des catégories depuis les fichiers Excel
Pour préparer le Bloc 4 - Création des catégories dans Odoo
"""

import os
from openpyxl import load_workbook
from collections import defaultdict

# Chemins des fichiers
BASE_PATH = "/Users/mhoar/Desktop/GEL OI/donnees"
FILES = {
    'produits': os.path.join(BASE_PATH, "7-Liste des Produits.xlsx"),
    'clients': os.path.join(BASE_PATH, "8-Liste des Clients.xlsx"),
    'fournisseurs': os.path.join(BASE_PATH, "9-Liste des fournisseurs.xlsx"),
}

def get_column_values(ws, col_name, max_rows=5000):
    """Récupère toutes les valeurs uniques d'une colonne"""
    values = set()
    header_row = None
    col_idx = None
    
    # Trouver l'en-tête
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        for col, cell in enumerate(row, 1):
            if cell and col_name.lower() in str(cell).lower():
                header_row = row_idx
                col_idx = col
                break
        if col_idx:
            break
    
    if not col_idx:
        return None, []
    
    # Récupérer les valeurs
    for row in ws.iter_rows(min_row=header_row+1, max_row=max_rows, min_col=col_idx, max_col=col_idx, values_only=True):
        if row[0] and str(row[0]).strip():
            values.add(str(row[0]).strip())
    
    return col_idx, sorted(values)

def analyze_excel_structure(filepath, name):
    """Analyse la structure d'un fichier Excel"""
    print(f"\n{'='*60}")
    print(f"📄 Analyse: {name}")
    print(f"   Fichier: {filepath}")
    print('='*60)
    
    if not os.path.exists(filepath):
        print(f"   ❌ Fichier non trouvé!")
        return {}
    
    wb = load_workbook(filepath, read_only=True, data_only=True)
    results = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n   📑 Feuille: {sheet_name}")
        
        # Lire les en-têtes (premières lignes)
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            if any(row):
                headers = [str(c) if c else '' for c in row]
                break
        
        print(f"   Colonnes: {len(headers)}")
        
        # Afficher les colonnes
        for i, h in enumerate(headers[:30], 1):
            if h:
                print(f"      [{i:2}] {h}")
        
        results[sheet_name] = headers
    
    wb.close()
    return results

def extract_categories():
    """Extrait toutes les catégories des fichiers"""
    
    categories = {
        'produits': defaultdict(set),
        'clients': defaultdict(set),
        'fournisseurs': defaultdict(set),
    }
    
    # =========================================================================
    # PRODUITS
    # =========================================================================
    print("\n" + "="*70)
    print("🔍 EXTRACTION DES CATÉGORIES - PRODUITS")
    print("="*70)
    
    if os.path.exists(FILES['produits']):
        wb = load_workbook(FILES['produits'], read_only=True, data_only=True)
        ws = wb.active
        
        # Colonnes potentielles pour les catégories
        category_columns = [
            'categorie', 'catégorie', 'category', 'famille', 'sous-famille',
            'sous famille', 'type', 'segment', 'gamme', 'rayon', 'univers',
            'nature', 'groupe', 'class', 'marque', 'origine', 'zone'
        ]
        
        # Lire les en-têtes
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            if any(row):
                headers = [str(c).lower() if c else '' for c in row]
                break
        
        print(f"\nColonnes trouvées: {len(headers)}")
        
        # Extraire les valeurs des colonnes de catégories
        for col_idx, header in enumerate(headers, 1):
            for cat_col in category_columns:
                if cat_col in header:
                    values = set()
                    for row in ws.iter_rows(min_row=2, max_row=5000, min_col=col_idx, max_col=col_idx, values_only=True):
                        if row[0] and str(row[0]).strip():
                            values.add(str(row[0]).strip())
                    
                    if values:
                        categories['produits'][header] = values
                        print(f"\n   📦 {header.upper()} ({len(values)} valeurs)")
                        for v in sorted(values)[:20]:
                            print(f"      - {v}")
                        if len(values) > 20:
                            print(f"      ... et {len(values)-20} autres")
                    break
        
        wb.close()
    
    # =========================================================================
    # CLIENTS
    # =========================================================================
    print("\n" + "="*70)
    print("🔍 EXTRACTION DES CATÉGORIES - CLIENTS")
    print("="*70)
    
    if os.path.exists(FILES['clients']):
        wb = load_workbook(FILES['clients'], read_only=True, data_only=True)
        ws = wb.active
        
        category_columns = [
            'categorie', 'catégorie', 'type', 'groupe', 'enseigne', 'typologie',
            'secteur', 'canal', 'segment', 'zone', 'region', 'région', 'tag',
            'etiquette', 'classification', 'tarif', 'grille'
        ]
        
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            if any(row):
                headers = [str(c).lower() if c else '' for c in row]
                break
        
        print(f"\nColonnes trouvées: {len(headers)}")
        
        for col_idx, header in enumerate(headers, 1):
            for cat_col in category_columns:
                if cat_col in header:
                    values = set()
                    for row in ws.iter_rows(min_row=2, max_row=5000, min_col=col_idx, max_col=col_idx, values_only=True):
                        if row[0] and str(row[0]).strip():
                            values.add(str(row[0]).strip())
                    
                    if values:
                        categories['clients'][header] = values
                        print(f"\n   👥 {header.upper()} ({len(values)} valeurs)")
                        for v in sorted(values)[:20]:
                            print(f"      - {v}")
                        if len(values) > 20:
                            print(f"      ... et {len(values)-20} autres")
                    break
        
        wb.close()
    
    # =========================================================================
    # FOURNISSEURS
    # =========================================================================
    print("\n" + "="*70)
    print("🔍 EXTRACTION DES CATÉGORIES - FOURNISSEURS")
    print("="*70)
    
    if os.path.exists(FILES['fournisseurs']):
        wb = load_workbook(FILES['fournisseurs'], read_only=True, data_only=True)
        ws = wb.active
        
        category_columns = [
            'categorie', 'catégorie', 'type', 'groupe', 'zone', 'pays', 'region',
            'région', 'origine', 'incoterm', 'transitaire'
        ]
        
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            if any(row):
                headers = [str(c).lower() if c else '' for c in row]
                break
        
        print(f"\nColonnes trouvées: {len(headers)}")
        
        for col_idx, header in enumerate(headers, 1):
            for cat_col in category_columns:
                if cat_col in header:
                    values = set()
                    for row in ws.iter_rows(min_row=2, max_row=5000, min_col=col_idx, max_col=col_idx, values_only=True):
                        if row[0] and str(row[0]).strip():
                            values.add(str(row[0]).strip())
                    
                    if values:
                        categories['fournisseurs'][header] = values
                        print(f"\n   🏭 {header.upper()} ({len(values)} valeurs)")
                        for v in sorted(values)[:20]:
                            print(f"      - {v}")
                        if len(values) > 20:
                            print(f"      ... et {len(values)-20} autres")
                    break
        
        wb.close()
    
    return categories

def main():
    print("╔══════════════════════════════════════════════════════════════════╗")
    print("║     EXTRACTION DES CATÉGORIES POUR BLOC 4 - ODOO                 ║")
    print("╚══════════════════════════════════════════════════════════════════╝")
    
    # Vérifier les fichiers
    print("\n📂 Vérification des fichiers...")
    for name, path in FILES.items():
        exists = "✅" if os.path.exists(path) else "❌"
        print(f"   {exists} {name}: {path}")
    
    # Analyser la structure
    print("\n" + "="*70)
    print("📊 ANALYSE DE LA STRUCTURE DES FICHIERS")
    print("="*70)
    
    for name, path in FILES.items():
        analyze_excel_structure(path, name)
    
    # Extraire les catégories
    categories = extract_categories()
    
    # Résumé
    print("\n" + "="*70)
    print("📋 RÉSUMÉ DES CATÉGORIES À CRÉER DANS ODOO")
    print("="*70)
    
    print("\n┌─────────────────────────────────────────────────────────────────┐")
    print("│ CATÉGORIES PRODUITS (product.category)                          │")
    print("└─────────────────────────────────────────────────────────────────┘")
    for col, values in categories['produits'].items():
        print(f"\n   {col}: {len(values)} valeurs")
        for v in sorted(values):
            print(f"      • {v}")
    
    print("\n┌─────────────────────────────────────────────────────────────────┐")
    print("│ ÉTIQUETTES CLIENTS (res.partner.category)                       │")
    print("└─────────────────────────────────────────────────────────────────┘")
    for col, values in categories['clients'].items():
        print(f"\n   {col}: {len(values)} valeurs")
        for v in sorted(values):
            print(f"      • {v}")
    
    print("\n┌─────────────────────────────────────────────────────────────────┐")
    print("│ GROUPES FOURNISSEURS (champ x_groupe_fournisseur)               │")
    print("└─────────────────────────────────────────────────────────────────┘")
    for col, values in categories['fournisseurs'].items():
        print(f"\n   {col}: {len(values)} valeurs")
        for v in sorted(values):
            print(f"      • {v}")
    
    print("\n" + "="*70)
    print("✅ Extraction terminée!")
    print("="*70)

if __name__ == "__main__":
    main()
