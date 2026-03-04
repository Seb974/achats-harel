#!/usr/bin/env python3
"""
Import du plan comptable Cegid vers Odoo - AH CHOU
"""

import logging
from openpyxl import load_workbook
from odoo_client import OdooClient
from config import PLAN_COMPTABLE_FILE, IMPORT_LIMIT, DRY_RUN

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class AccountImporter:
    """Importeur du plan comptable depuis Excel vers Odoo"""
    
    # Ligne d'en-tête dans le fichier Excel
    HEADER_ROW = 6  # Les données commencent à la ligne 7 (index 6)
    
    # Colonnes (basé sur l'analyse du fichier)
    COLUMNS = {
        'numero': 1,     # Numéro de compte
        'intitule': 2,   # Intitulé
        'type': 3,       # Type (C = Client, F = Fournisseur)
        'cle': 4,        # Clé
        'collectif': 5,  # Collectif
        'tel1': 6,       # Tel1
        'fax': 7,        # Fax
        'ville': 8,      # Ville
    }
    
    def __init__(self):
        self.odoo = OdooClient()
        self.stats = {
            'total': 0,
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': 0,
        }
    
    def get_cell_value(self, row, col_name):
        """Récupère la valeur d'une cellule"""
        idx = self.COLUMNS.get(col_name)
        if idx is None or idx >= len(row):
            return None
        value = row[idx]
        if value is None:
            return None
        if isinstance(value, str):
            return value.strip() if value.strip() else None
        return value
    
    def get_account_type(self, code, type_field=None):
        """
        Détermine le type de compte selon le PCG français
        
        Odoo 17+ utilise ces types:
        - equity (capitaux propres)
        - asset_non_current (actifs non courants)
        - asset_current (actifs courants)
        - asset_receivable (créances clients)
        - asset_cash (trésorerie)
        - liability_non_current (dettes long terme)
        - liability_current (dettes court terme)
        - liability_payable (dettes fournisseurs)
        - income (produits)
        - income_other (autres produits)
        - expense (charges)
        - expense_depreciation (amortissements)
        - expense_direct_cost (coûts directs)
        - off_balance (hors bilan)
        """
        
        if not code:
            return 'off_balance'
        
        code = str(code).strip()
        
        # Comptes auxiliaires (01xx, 08xx) - à ignorer
        if code.startswith('01') or code.startswith('08'):
            return None  # Signale qu'il faut ignorer
        
        # Classe 1 - Capitaux propres
        if code.startswith('1'):
            return 'equity'
        
        # Classe 2 - Immobilisations
        if code.startswith('2'):
            return 'asset_non_current'
        
        # Classe 3 - Stocks
        if code.startswith('3'):
            return 'asset_current'
        
        # Classe 4 - Tiers
        if code.startswith('4'):
            # 40x - Fournisseurs
            if code.startswith('40'):
                return 'liability_payable'
            # 41x - Clients
            if code.startswith('41'):
                return 'asset_receivable'
            # 42x - Personnel
            if code.startswith('42'):
                return 'liability_current'
            # 43x - Sécurité sociale
            if code.startswith('43'):
                return 'liability_current'
            # 44x - État (TVA, etc.)
            if code.startswith('44'):
                return 'liability_current'
            # 45x - Groupe et associés
            if code.startswith('45'):
                return 'liability_current'
            # 46x - Débiteurs/créditeurs divers
            if code.startswith('46'):
                return 'asset_current'
            # 47x - Comptes transitoires
            if code.startswith('47'):
                return 'liability_current'
            # 48x - Provisions
            if code.startswith('48'):
                return 'liability_current'
            # 49x - Dépréciations
            if code.startswith('49'):
                return 'expense_depreciation'
            return 'liability_current'
        
        # Classe 5 - Trésorerie
        if code.startswith('5'):
            # 51x - Banques
            if code.startswith('51'):
                return 'asset_cash'
            # 53x - Caisse
            if code.startswith('53'):
                return 'asset_cash'
            return 'asset_cash'
        
        # Classe 6 - Charges
        if code.startswith('6'):
            # 60x - Achats
            if code.startswith('60'):
                return 'expense_direct_cost'
            # 68x - Dotations aux amortissements
            if code.startswith('68'):
                return 'expense_depreciation'
            return 'expense'
        
        # Classe 7 - Produits
        if code.startswith('7'):
            # 70x - Ventes
            if code.startswith('70'):
                return 'income'
            # 78x - Reprises
            if code.startswith('78'):
                return 'income_other'
            return 'income'
        
        # Classe 8 - Comptes spéciaux
        if code.startswith('8'):
            return 'off_balance'
        
        return 'off_balance'
    
    def row_to_account_values(self, row):
        """Convertit une ligne Excel en valeurs account.account"""
        
        code = self.get_cell_value(row, 'numero')
        if not code:
            return None
        
        code = str(code).strip()
        
        # Ignorer les comptes auxiliaires
        if code.startswith('01') or code.startswith('08'):
            return None
        
        intitule = self.get_cell_value(row, 'intitule')
        if not intitule:
            intitule = code
        
        account_type = self.get_account_type(code)
        if account_type is None:
            return None
        
        values = {
            'code': code,
            'name': intitule,
            'account_type': account_type,
        }
        
        # Les comptes de tiers sont lettrables
        if account_type in ['asset_receivable', 'liability_payable']:
            values['reconcile'] = True
        
        return values
    
    def import_accounts(self):
        """Importe le plan comptable depuis Excel"""
        logger.info(f"Chargement du fichier Excel: {PLAN_COMPTABLE_FILE}")
        
        try:
            wb = load_workbook(PLAN_COMPTABLE_FILE, read_only=True)
        except FileNotFoundError:
            logger.error(f"Fichier non trouvé: {PLAN_COMPTABLE_FILE}")
            return
        
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        
        # Skip les lignes d'en-tête (les 6 premières lignes)
        data_rows = rows[self.HEADER_ROW:]
        
        if IMPORT_LIMIT:
            data_rows = data_rows[:IMPORT_LIMIT]
        
        self.stats['total'] = len(data_rows)
        logger.info(f"Traitement de {self.stats['total']} lignes...")
        
        for i, row in enumerate(data_rows, 1):
            try:
                values = self.row_to_account_values(row)
                
                if not values:
                    self.stats['skipped'] += 1
                    continue
                
                # Chercher par code
                code = values.get('code')
                existing = self.odoo.search(
                    'account.account',
                    [('code', '=', code)],
                    limit=1
                )
                
                if existing:
                    # Mise à jour
                    self.odoo.write('account.account', existing, values)
                    self.stats['updated'] += 1
                    logger.debug(f"[{i}] MAJ: {code} - {values['name']}")
                else:
                    # Création
                    self.odoo.create('account.account', values)
                    self.stats['created'] += 1
                    logger.debug(f"[{i}] Créé: {code} - {values['name']}")
                
                # Progress
                if i % 100 == 0:
                    logger.info(f"Progression: {i}/{self.stats['total']}")
                    
            except Exception as e:
                self.stats['errors'] += 1
                logger.error(f"Erreur ligne {i}: {e}")
        
        wb.close()
        self.print_stats()
    
    def print_stats(self):
        """Affiche les statistiques d'import"""
        print("\n" + "="*50)
        print("RÉSULTATS IMPORT PLAN COMPTABLE")
        print("="*50)
        print(f"  Total lignes     : {self.stats['total']}")
        print(f"  Créés            : {self.stats['created']}")
        print(f"  Mis à jour       : {self.stats['updated']}")
        print(f"  Ignorés (aux.)   : {self.stats['skipped']}")
        print(f"  Erreurs          : {self.stats['errors']}")
        print("="*50)
        if DRY_RUN:
            print("⚠️  MODE DRY-RUN: Aucune modification réelle")


def main():
    """Point d'entrée principal"""
    print("="*50)
    print("IMPORT PLAN COMPTABLE - CEGID vers ODOO")
    print("="*50)
    
    if DRY_RUN:
        print("⚠️  Mode DRY-RUN activé (pas d'écriture)")
    
    importer = AccountImporter()
    importer.import_accounts()


if __name__ == '__main__':
    main()
